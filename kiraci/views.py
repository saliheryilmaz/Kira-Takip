from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.utils import timezone
from django.db.models import Sum
from decimal import Decimal
from .models import Kiraci, Odeme, UserProfile
from .forms import KiraciForm, OdemeForm, KayitForm, ProfilForm


# ─── Auth ───

def giris(request):
    if request.user.is_authenticated:
        return redirect('kiraci_listesi')
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('kiraci_listesi')
        messages.error(request, 'Kullanıcı adı veya şifre hatalı.')
    return render(request, 'kiraci/giris.html')


def kayit(request):
    if request.user.is_authenticated:
        return redirect('kiraci_listesi')
    if request.method == 'POST':
        form = KayitForm(request.POST)
        if form.is_valid():
            user = form.save()
            UserProfile.objects.create(user=user)
            login(request, user)
            messages.success(request, 'Hesabınız oluşturuldu.')
            return redirect('kiraci_listesi')
    else:
        form = KayitForm()
    return render(request, 'kiraci/kayit.html', {'form': form})


@login_required
def cikis(request):
    logout(request)
    return redirect('giris')


@login_required
def profil(request):
    profil_obj, _ = UserProfile.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        form = ProfilForm(request.POST, instance=profil_obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profil güncellendi.')
            return redirect('profil')
    else:
        form = ProfilForm(instance=profil_obj)
    return render(request, 'kiraci/profil.html', {'form': form})


# ─── Kiracı ───

@login_required
def kiraci_listesi(request):
    kiraciler = Kiraci.objects.filter(user=request.user).order_by('-aktif', 'firma_adi')
    bugun = timezone.now().date()
    aktif_qs = kiraciler.filter(aktif=True)
    aktif_sayisi = aktif_qs.count()
    toplam_beklenen = sum(k.toplam_beklenen() for k in aktif_qs)
    toplam_odenen = sum(k.toplam_odenen() for k in aktif_qs)
    toplam_borc = sum(k.toplam_borc() for k in aktif_qs)
    bu_ay_odeme = Odeme.objects.filter(
        kiraci__user=request.user, yil=bugun.year, ay=bugun.month
    ).values_list('kiraci_id', flat=True)
    bu_ay_odemeyen = aktif_qs.exclude(id__in=bu_ay_odeme)
    context = {
        'kiraciler': kiraciler,
        'aktif_sayisi': aktif_sayisi,
        'toplam_beklenen': toplam_beklenen,
        'toplam_odenen': toplam_odenen,
        'toplam_borc': toplam_borc,
        'bu_ay_odemeyen': bu_ay_odemeyen,
        'bugun': bugun,
    }
    return render(request, 'kiraci/kiraci_listesi.html', context)


@login_required
def kiraci_detay(request, pk):
    kiraci = get_object_or_404(Kiraci, pk=pk, user=request.user)
    ay_listesi = kiraci.ay_listesi()
    context = {
        'kiraci': kiraci,
        'ay_listesi': ay_listesi,
        'toplam_beklenen': kiraci.toplam_beklenen(),
        'toplam_odenen': kiraci.toplam_odenen(),
        'toplam_borc': kiraci.toplam_borc(),
    }
    return render(request, 'kiraci/kiraci_detay.html', context)


@login_required
def kiraci_ekle(request):
    if request.method == 'POST':
        form = KiraciForm(request.POST, request.FILES)
        if form.is_valid():
            kiraci = form.save(commit=False)
            kiraci.user = request.user
            kiraci.save()
            messages.success(request, f'"{kiraci.firma_adi}" başarıyla eklendi.')
            return redirect('kiraci_detay', pk=kiraci.pk)
    else:
        form = KiraciForm()
    return render(request, 'kiraci/kiraci_form.html', {'form': form, 'baslik': 'Yeni Kiracı Ekle'})


@login_required
def kiraci_duzenle(request, pk):
    kiraci = get_object_or_404(Kiraci, pk=pk, user=request.user)
    if request.method == 'POST':
        form = KiraciForm(request.POST, request.FILES, instance=kiraci)
        if form.is_valid():
            form.save()
            messages.success(request, 'Kiracı bilgileri güncellendi.')
            return redirect('kiraci_detay', pk=kiraci.pk)
    else:
        form = KiraciForm(instance=kiraci)
    return render(request, 'kiraci/kiraci_form.html', {
        'form': form, 'kiraci': kiraci,
        'baslik': f'{kiraci.firma_adi} — Düzenle'
    })


@login_required
def kiraci_sil(request, pk):
    kiraci = get_object_or_404(Kiraci, pk=pk, user=request.user)
    if request.method == 'POST':
        firma_adi = kiraci.firma_adi
        kiraci.delete()
        messages.success(request, f'"{firma_adi}" silindi.')
        return redirect('kiraci_listesi')
    return render(request, 'kiraci/kiraci_sil_onay.html', {'kiraci': kiraci})


# ─── Ödeme ───

@login_required
def odeme_ekle(request, kiraci_pk):
    kiraci = get_object_or_404(Kiraci, pk=kiraci_pk, user=request.user)
    if request.method == 'POST':
        form = OdemeForm(request.POST, kiraci=kiraci)
        if form.is_valid():
            odeme = form.save(commit=False)
            odeme.kiraci = kiraci
            odeme.save()
            messages.success(request, f'{odeme.ay_adi} {odeme.yil} ödemesi kaydedildi.')
            return redirect('kiraci_detay', pk=kiraci.pk)
    else:
        initial = {}
        if 'yil' in request.GET:
            initial['yil'] = request.GET['yil']
        if 'ay' in request.GET:
            initial['ay'] = request.GET['ay']
        form = OdemeForm(kiraci=kiraci, initial=initial)
    return render(request, 'kiraci/odeme_form.html', {'form': form, 'kiraci': kiraci})


@login_required
def odeme_duzenle(request, pk):
    odeme = get_object_or_404(Odeme, pk=pk, kiraci__user=request.user)
    kiraci = odeme.kiraci
    if request.method == 'POST':
        form = OdemeForm(request.POST, instance=odeme, kiraci=kiraci)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ödeme güncellendi.')
            return redirect('kiraci_detay', pk=kiraci.pk)
    else:
        form = OdemeForm(instance=odeme, kiraci=kiraci)
    return render(request, 'kiraci/odeme_form.html', {
        'form': form, 'kiraci': kiraci, 'odeme': odeme, 'duzenleme': True
    })


@login_required
def odeme_sil(request, pk):
    odeme = get_object_or_404(Odeme, pk=pk, kiraci__user=request.user)
    kiraci = odeme.kiraci
    if request.method == 'POST':
        odeme.delete()
        messages.success(request, 'Ödeme kaydı silindi.')
        return redirect('kiraci_detay', pk=kiraci.pk)
    return render(request, 'kiraci/odeme_sil_onay.html', {'odeme': odeme, 'kiraci': kiraci})


# ─── Aylık Özet ───

@login_required
def aylik_ozet(request):
    bugun = timezone.now().date()
    yil = int(request.GET.get('yil', bugun.year))
    ay = int(request.GET.get('ay', bugun.month))
    ay_adlari = {1:'Ocak',2:'Şubat',3:'Mart',4:'Nisan',5:'Mayıs',6:'Haziran',
                 7:'Temmuz',8:'Ağustos',9:'Eylül',10:'Ekim',11:'Kasım',12:'Aralık'}
    kiraciler = Kiraci.objects.filter(
        user=request.user, aktif=True,
        kira_baslangic_tarihi__lte=f'{yil}-{ay:02d}-01'
    )
    ozet_listesi = []
    toplam_beklenen = Decimal('0')
    toplam_odenen = Decimal('0')
    for kiraci in kiraciler:
        odemeler = kiraci.odemeler.filter(yil=yil, ay=ay)
        beklenen = kiraci.aylik_kira_tutari
        odenen = odemeler.aggregate(t=Sum('odenen_tutar'))['t'] or Decimal('0')
        toplam_beklenen += beklenen
        toplam_odenen += odenen
        ozet_listesi.append({
            'kiraci': kiraci, 'odemeler': odemeler,
            'beklenen': beklenen, 'odenen': odenen,
            'eksik': beklenen - odenen if odenen < beklenen else Decimal('0'),
            'fazla': odenen - beklenen if odenen > beklenen else Decimal('0'),
        })
    onceki_yil, onceki_ay = (yil - 1, 12) if ay == 1 else (yil, ay - 1)
    sonraki_yil, sonraki_ay = (yil + 1, 1) if ay == 12 else (yil, ay + 1)
    context = {
        'ozet_listesi': ozet_listesi, 'yil': yil, 'ay': ay,
        'ay_adi': ay_adlari[ay], 'toplam_beklenen': toplam_beklenen,
        'toplam_odenen': toplam_odenen,
        'toplam_eksik': toplam_beklenen - toplam_odenen,
        'onceki_yil': onceki_yil, 'onceki_ay': onceki_ay,
        'sonraki_yil': sonraki_yil, 'sonraki_ay': sonraki_ay,
        'ay_adlari': ay_adlari, 'yillar': range(2020, bugun.year + 2),
    }
    return render(request, 'kiraci/aylik_ozet.html', context)


# ─── Bildirim Yanıt ───

from .models import BildirimLog

def odeme_onayla(request, token):
    """WhatsApp/Email'deki 'Evet Alındı' linkine tıklanınca çalışır."""
    log = get_object_or_404(BildirimLog, token=token)

    if log.yanitlandi:
        return render(request, 'kiraci/bildirim_yanit.html', {
            'mesaj': 'Bu bildirim zaten yanıtlanmış.',
            'tur': 'bilgi'
        })

    kiraci = log.kiraci
    # Ödeme zaten var mı?
    if not Odeme.objects.filter(kiraci=kiraci, yil=log.yil, ay=log.ay).exists():
        Odeme.objects.create(
            kiraci=kiraci,
            yil=log.yil,
            ay=log.ay,
            odenen_tutar=kiraci.aylik_kira_tutari,
            odeme_turu='elden',
            aciklama='WhatsApp/Email bildirimi üzerinden onaylandı.',
        )

    log.yanitlandi = True
    log.save()

    return render(request, 'kiraci/bildirim_yanit.html', {
        'mesaj': f'✅ {kiraci.firma_adi} — {log.ay}/{log.yil} ödemesi alındı olarak kaydedildi.',
        'tur': 'basari'
    })


def odeme_reddet(request, token):
    """'Hayır' linkine tıklanınca — yarın tekrar bildirim gidecek."""
    log = get_object_or_404(BildirimLog, token=token)
    log.yanitlandi = True
    log.save()

    return render(request, 'kiraci/bildirim_yanit.html', {
        'mesaj': f'Anlaşıldı. {log.kiraci.firma_adi} için yarın tekrar hatırlatılacak.',
        'tur': 'bilgi'
    })


# ─── Excel Export ───

import io
from django.http import HttpResponse

@login_required
def kiraci_excel(request, pk):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl kurulu değil.', status=500)

    kiraci = get_object_or_404(Kiraci, pk=pk, user=request.user)
    ay_listesi = kiraci.ay_listesi()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = kiraci.firma_adi[:30]

    # Renkler
    yesil = PatternFill('solid', fgColor='00C896')
    koyu  = PatternFill('solid', fgColor='111418')
    gri   = PatternFill('solid', fgColor='1F2630')
    kirmizi_fill = PatternFill('solid', fgColor='FF4757')
    sari_fill    = PatternFill('solid', fgColor='FFA502')

    bold_white = Font(bold=True, color='FFFFFF', size=11)
    bold_black = Font(bold=True, color='000000', size=11)
    normal     = Font(color='E8EDF2', size=10)
    green_font = Font(color='00C896', size=10, bold=True)
    red_font   = Font(color='FF4757', size=10, bold=True)
    yellow_font= Font(color='FFA502', size=10)

    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')
    right  = Alignment(horizontal='right',  vertical='center')

    thin = Side(style='thin', color='2A3340')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Başlık ──
    ws.merge_cells('A1:G1')
    c = ws['A1']
    c.value = kiraci.firma_adi
    c.font = Font(bold=True, color='000000', size=14)
    c.fill = yesil
    c.alignment = center
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:G2')
    c = ws['A2']
    c.value = f"Aylık Kira: {int(kiraci.aylik_kira_tutari):,} ₺  |  Başlangıç: {kiraci.kira_baslangic_tarihi.strftime('%d.%m.%Y')}  |  Depozit: {int(kiraci.depozit):,} ₺"
    c.font = Font(color='8A9AB0', size=10)
    c.fill = koyu
    c.alignment = center
    ws.row_dimensions[2].height = 18

    # ── Sütun başlıkları ──
    basliklar = ['Dönem', 'Beklenen (₺)', 'Ödenen (₺)', 'Eksik (₺)', 'Ödeme Türü', 'Tarih', 'Açıklama']
    ws.append([])  # boş satır
    ws.row_dimensions[3].height = 6

    row = 4
    for i, b in enumerate(basliklar, 1):
        c = ws.cell(row=row, column=i, value=b)
        c.font = bold_white
        c.fill = gri
        c.alignment = center
        c.border = border
    ws.row_dimensions[row].height = 22

    # ── Veriler ──
    for item in reversed(ay_listesi):  # eskiden yeniye
        donem = f"{item['ay_adi']} {item['yil']}"
        beklenen = int(item['beklenen'])
        odenen_toplam = int(item['odenen'])
        eksik = int(item['eksik'])

        if not item['odemeler']:
            # Ödeme yok — tek satır
            row += 1
            satirlar = [
                (donem, beklenen, 0, beklenen, '—', '—', '—')
            ]
            for i, val in enumerate(satirlar[0], 1):
                c = ws.cell(row=row, column=i, value=val)
                c.font = red_font if i in (3, 4) else normal
                c.fill = PatternFill('solid', fgColor='1A0A0D')
                c.alignment = right if i in (2, 3, 4) else left
                c.border = border
            ws.row_dimensions[row].height = 18
        else:
            for j, odeme in enumerate(item['odemeler']):
                row += 1
                tur_map = {'resmi': '🏦 Resmi', 'elden': '💵 Elden', 'gayri_resmi': '📝 Gayri Resmi'}
                tur = tur_map.get(odeme.odeme_turu, odeme.odeme_turu)
                tarih = odeme.odeme_tarihi.strftime('%d.%m.%Y')
                aciklama = odeme.aciklama or ''

                if j == 0:
                    d_val = donem
                    b_val = beklenen
                    e_val = eksik if eksik > 0 else 0
                else:
                    d_val = ''
                    b_val = ''
                    e_val = ''

                satirlar = [d_val, b_val, int(odeme.odenen_tutar), e_val, tur, tarih, aciklama]
                fill_color = PatternFill('solid', fgColor='0D1A14') if odenen_toplam >= beklenen else PatternFill('solid', fgColor='1A1200')

                for i, val in enumerate(satirlar, 1):
                    c = ws.cell(row=row, column=i, value=val)
                    if i == 3:
                        c.font = green_font
                    elif i == 4 and val and val > 0:
                        c.font = red_font
                    elif i == 4 and (val == 0 or val == ''):
                        c.font = Font(color='00C896', size=10)
                    else:
                        c.font = normal
                    c.fill = fill_color
                    c.alignment = right if i in (2, 3, 4) else left
                    c.border = border
                ws.row_dimensions[row].height = 18

    # ── Özet satırı ──
    row += 2
    toplam_beklenen = int(kiraci.toplam_beklenen())
    toplam_odenen   = int(kiraci.toplam_odenen())
    toplam_borc     = int(kiraci.toplam_borc())

    ozet = [('TOPLAM BEKLENEN', toplam_beklenen), ('TOPLAM ÖDENEN', toplam_odenen), ('TOPLAM ALACAK', toplam_borc)]
    for baslik, deger in ozet:
        ws.cell(row=row, column=1, value=baslik).font = Font(bold=True, color='8A9AB0', size=10)
        ws.cell(row=row, column=1).fill = koyu
        ws.cell(row=row, column=1).alignment = left
        c = ws.cell(row=row, column=2, value=deger)
        c.font = Font(bold=True, color='00C896' if baslik != 'TOPLAM ALACAK' or deger == 0 else 'FF4757', size=11)
        c.fill = koyu
        c.alignment = right
        ws.row_dimensions[row].height = 20
        row += 1

    # ── Sütun genişlikleri ──
    genislikler = [18, 16, 16, 14, 18, 14, 30]
    for i, g in enumerate(genislikler, 1):
        ws.column_dimensions[get_column_letter(i)].width = g

    # ── İndir ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    dosya_adi = f"{kiraci.firma_adi.replace(' ', '_')}_odeme_gecmisi.xlsx"
    response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{dosya_adi}"'
    return response
